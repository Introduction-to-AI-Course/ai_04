# -*- coding:utf8 -*-
from openpyxl.reader.excel import load_workbook
from stanfordcorenlp import StanfordCoreNLP
from polite import *
from thx import *
from tel import *
from fromwhom import *
from date import *
import jieba
import pandas as pd
from gensim import corpora, models, similarities

nlp = StanfordCoreNLP(r'E:\operation\stanford-corenlp-full-2020-04-20', lang='zh')  # 用于实体识别

score = ['N']
politeScore = ['I']
whoScore = ['M']
dateScore = ['K']
telScore = ['L']
thxScore = ['J']


# 数据预处理，除去多余字符
def preTreat(sentence):
    # sentence2 = str(sentence).replace('\t', '')
    # sentence2 = str(sentence2).replace('\n', '')
    sentence2 = str(sentence).replace('　', ' ')
    sentence2 = str(sentence2).replace('  ', ' ')
    sentence2 = str(sentence2).replace('，',',')
    sentence2 = str(sentence2).replace('。','.')
    sentence2 = str(sentence2).replace('、',',')
    sentence2 = str(sentence2).replace('？','?')
    sentence2 = str(sentence2).replace('！','!')
    # sentence2 = str(sentence2).replace('','')

    return sentence2

def WriteToExcel(inFile='infile/附件4.xlsx', outFile='outfile/评分后附件4.xlsx'):
    def writeOneCol(dataList, title='标题', TinFile=inFile, ToutFile=outFile):
        wb = load_workbook(TinFile)
        ws = wb['Sheet1']
        ws[dataList[0] + '1'] = title
        for i in range(1, len(dataList)):
            curRow = dataList[0] + str(i + 1)
            ws[curRow] = dataList[i]
        wb.save(ToutFile)

    writeOneCol(score, '答复总分', inFile, outFile)
    writeOneCol(telScore, '电话分', outFile, outFile)
    writeOneCol(thxScore, '感谢分', outFile, outFile)
    writeOneCol(politeScore, '礼貌分', outFile, outFile)
    writeOneCol(whoScore, '机构分', outFile, outFile)
    writeOneCol(dateScore, '日期分', outFile, outFile)

if __name__ == '__main__':
    t = 0
    path = r'infile/附件4.xlsx'
    out = r'outfile/问题3 评价完毕表格.xlsx'
    data = pd.read_excel(path)
    for i in data['答复意见']:
        t+=1
        print('第一阶段评价')
        print(str(t) + '====='*10)
        sentence = preTreat(i)
        politeScore.append(PoliteMark(sentence))
        dateScore.append(ReplyDate(sentence))
        telScore.append(Tel(sentence))
        thxScore.append(thx(sentence))
    t = 0
    for i in data['答复意见']:
        t+=1
        print('第二阶段评价')
        print(str(t) + '=====' * 10)
        sentence = preTreat(i)
        whoScore.append(FromWhom(sentence,nlp))
        score.append(politeScore[t]+dateScore[t]+whoScore[t]+telScore[t]+thxScore[t])
    WriteToExcel(path,out)

    data = pd.read_excel(out)
    data['留言时间'] = pd.to_datetime(data['留言时间'])
    data['答复时间'] = pd.to_datetime(data['答复时间'])

    # 可解释性
    expainarray = []
    for index, row in data.iterrows():
        sum = 0
        for x in row["答复意见"]:
            if x == '《':
                sum = sum + 1
        expainarray.append(sum)
    totoal = 0
    for x in expainarray:
        totoal = totoal + x
    data["可解释性"] = [x / (totoal / len(expainarray)) for x in expainarray]

    # 相关性
    datax = data['留言详情'].apply(lambda x: re.sub('x', ' ', x))
    datay = data['答复意见'].apply(lambda x: re.sub('x', ' ', x))
    jieba.load_userdict(r'data\newdic1.txt')
    data_cut_1 = datax.apply(lambda x: jieba.lcut(x))
    data_cut_2 = datay.apply(lambda x: jieba.lcut(x))
    stopWords = pd.read_csv(r'data\stopword.txt', encoding='GB18030', sep='hahah', engine='python')
    stopWords = ['≮', '≯', '≠', '≮', ' ', '会', '月', '日', '–', '！', '，', '。', '\u3000', '，', '\xa0', '\t', '\n'] + list(
        stopWords.iloc[:, 0])
    data_after_stop_1 = data_cut_1.apply(lambda x: [i for i in x if i not in stopWords])
    data_after_stop_2 = data_cut_2.apply(lambda x: [i for i in x if i not in stopWords])
    similararray = []
    for x in range(len(data)):
        texts = [data_after_stop_1[x], ['想要']]
        # 建立词典
        dictionary = corpora.Dictionary(texts)
        num_features = len(dictionary.token2id)
        # 基于词典，将【分词列表集】转换成【稀疏向量集】，称作【语料库】
        corpus = [dictionary.doc2bow(text) for text in texts]
        # 关键词转换为稀疏向量
        kw_vector = dictionary.doc2bow(data_after_stop_2[x])
        # 4、创建【TF-IDF模型】，传入【语料库】来训练
        tfidf = models.TfidfModel(corpus)
        # 5、用训练好的【TF-IDF模型】处理【被检索文本】和【搜索词】
        tf_texts = tfidf[corpus]
        tf_kw = tfidf[kw_vector]
        sparse_matrix = similarities.SparseMatrixSimilarity(tf_texts, num_features)
        similar = sparse_matrix[tf_kw]
        similararray.append(similar[0])
    data['相关性'] = [x for x in similararray]

    # 时间答复性
    time = []
    for index, row in data.iterrows():
        if (row['答复时间'] - row['留言时间']).days <= 3:
            time.append(5)
        elif (row['答复时间'] - row['留言时间']).days <= 7:
            time.append(4)
        elif (row['答复时间'] - row['留言时间']).days <= 14:
            time.append(3)
        elif (row['答复时间'] - row['留言时间']).days <= 30:
            time.append(2)
        elif (row['答复时间'] - row['留言时间']).days <= 90:
            time.append(1)
        else:
            time.append(0)
    data["回复时效性"] = [x for x in time]

    data.to_excel(out)